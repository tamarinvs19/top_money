from django.utils import timezone
from django.utils.timezone import make_aware
from datetime import datetime, timedelta
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db import models
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django import forms
import os
import shutil

from finance.models import Asset, Transaction, AssetType, TransactionType, WasteCategory, RefillCategory, BrokerageAccountType, get_asset_type_label
from finance.models import CashAsset, DebitCardAsset, DepositAsset, CreditCardAsset, BrokerageAsset, SavingAccount, EWalletAsset, Bank, BANKS, BankAsset, Provider, PROVIDERS, CashbackCategory
from finance.models import InvitationCode, BankCashbackCategory, BankCashbackMonth, BankCashbackSelection, CashbackCategory, BankCashbackMonthCategory
from finance.exchange_rate import ExchangeRateService


class SignupForm(UserCreationForm):
    invitation_code = forms.CharField(max_length=32, required=True, label='Invitation Code')
    
    def clean_invitation_code(self):
        code = self.cleaned_data.get('invitation_code', '').strip()
        if not code:
            raise forms.ValidationError('Invitation code is required.')
        try:
            invitation = InvitationCode.objects.get(code=code)
            if invitation.is_used:
                raise forms.ValidationError('This invitation code has already been used.')
        except InvitationCode.DoesNotExist:
            raise forms.ValidationError('Invalid invitation code.')
        return code


def signup(request):
    if request.method == 'POST':
        form = SignupForm(request.POST)
        if form.is_valid():
            user = form.save()
            code = InvitationCode.objects.get(code=form.cleaned_data['invitation_code'])
            code.used_by = user
            code.used_at = timezone.now()
            code.save()
            login(request, user)
            return redirect('transactions')
    else:
        form = SignupForm()
    return render(request, 'registration/signup.html', {'form': form})


@login_required
def transactions(request, year=None, month=None, asset_uuid=None):
    today = timezone.now()
    if year is None:
        year = today.year
    if month is None:
        month = today.month
    
    current_date = make_aware(datetime(int(year), int(month), 1))
    
    prev_month = current_date - timedelta(days=1)
    next_month = current_date + timedelta(days=31)
    next_month = next_month.replace(day=1)
    
    start_date = current_date
    end_date = current_date.replace(day=1, month=int(month) + 1) if int(month) < 12 else current_date.replace(year=int(year)+1, month=1)
    if int(month) == 12:
        end_date = make_aware(datetime(int(year) + 1, 1, 1)) - timedelta(seconds=1)
    else:
        from calendar import monthrange
        _, days = monthrange(int(year), int(month))
        end_date = make_aware(datetime(int(year), int(month), days, 23, 59, 59))
    
    transactions_list = Transaction.objects.filter(
        user=request.user,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('from_asset', 'to_asset').order_by('-date')
    
    if asset_uuid:
        transactions_list = transactions_list.filter(
            models.Q(from_asset_id=asset_uuid) | models.Q(to_asset_id=asset_uuid)
        )
    
    grouped = {}
    for t in transactions_list:
        day = t.date.date()
        if day not in grouped:
            grouped[day] = {'transactions': [], 'day_total': Decimal('0')}
        grouped[day]['transactions'].append(t)
        
        if t.type == TransactionType.REFILL:
            amount = t.amount
        elif t.type == TransactionType.WASTE:
            amount = -t.amount
        else:
            amount = Decimal('0')
        grouped[day]['day_total'] += amount
    
    month_income = Decimal('0')
    month_expense = Decimal('0')
    
    for day_data in grouped.values():
        for t in day_data['transactions']:
            if t.type == TransactionType.REFILL:
                month_income += t.amount
            elif t.type == TransactionType.WASTE:
                month_expense += t.amount
    
    month_balance = month_income - month_expense
    
    return render(request, 'transactions.html', {
        'transactions_by_day': grouped,
        'year': int(year),
        'month': int(month),
        'asset': Asset.objects.filter(id=asset_uuid).first() or "",
        'prev_month': prev_month,
        'next_month': next_month,
        'month_income': month_income,
        'month_expense': month_expense,
        'month_balance': month_balance,
    })


@login_required
def transaction_add(request, year=None, month=None, day=None):
    initial_date = None
    now = datetime.now()
    if year and month and day:
        initial_date = make_aware(datetime(int(year), int(month), int(day), now.hour, now.minute))
    elif year and month:
        initial_date = make_aware(datetime(int(year), int(month), 1, now.hour, now.minute))
    else:
        initial_date = make_aware(now)
    
    assets = Asset.objects.select_subclasses().filter(user=request.user, is_active=True)
    
    if request.method == 'POST':
        t_type = request.POST.get('type')
        amount = Decimal(request.POST.get('amount'))
        currency = request.POST.get('currency')
        category = request.POST.get('category') or ''
        description = request.POST.get('description')
        date_str = request.POST.get('date')
        time_str = request.POST.get('time', '00:00')
        date = make_aware(datetime.strptime(f'{date_str} {time_str}', '%Y-%m-%d %H:%M'))
        
        from_asset_id = request.POST.get('from_asset')
        to_asset_id = request.POST.get('to_asset')
        from_asset_rate = Decimal(request.POST.get('from_asset_rate', '1'))
        to_asset_rate = Decimal(request.POST.get('to_asset_rate', '1'))
        commission_rate = Decimal(request.POST.get('commission_rate', '0') or '0')
        commission_type = request.POST.get('commission_type', 'PERCENT')
        
        Transaction.objects.create(
            user=request.user,
            type=t_type,
            amount=amount,
            currency=currency,
            category=category,
            description=description,
            date=date,
            from_asset_id=from_asset_id if from_asset_id else None,
            from_asset_rate=from_asset_rate,
            to_asset_id=to_asset_id if to_asset_id else None,
            to_asset_rate=to_asset_rate,
            commission_rate=commission_rate,
            commission_type=commission_type,
        )
        
        return redirect('transactions')
    
    return render(request, 'transaction_form.html', {
        'assets': assets,
        'transaction_types': [t for t in TransactionType.choices if t[0] != TransactionType.CHANGING_BALANCE],
        'transaction_categories': WasteCategory.choices,
        'refill_categories': RefillCategory.choices,
        'initial_date': initial_date.strftime('%Y-%m-%d'),
        'initial_time': initial_date.strftime('%H:%M'),
    })


@login_required
def transaction_edit(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, user=request.user)
    assets = Asset.objects.select_subclasses().filter(user=request.user, is_active=True)
    
    if request.method == 'POST':
        transaction.type = request.POST.get('type')
        transaction.amount = Decimal(request.POST.get('amount'))
        transaction.currency = request.POST.get('currency')
        transaction.category = request.POST.get('category') or ''
        transaction.description = request.POST.get('description')
        transaction.date = make_aware(datetime.strptime(f"{request.POST.get('date')} {request.POST.get('time', '00:00')}", '%Y-%m-%d %H:%M'))
        
        from_asset_id = request.POST.get('from_asset')
        to_asset_id = request.POST.get('to_asset')
        from_asset_rate = Decimal(request.POST.get('from_asset_rate', '1'))
        to_asset_rate = Decimal(request.POST.get('to_asset_rate', '1'))
        commission_rate = Decimal(request.POST.get('commission_rate', '0') or '0')
        commission_type = request.POST.get('commission_type', 'PERCENT')
        
        transaction.from_asset_id = from_asset_id if from_asset_id else None
        transaction.from_asset_rate = from_asset_rate
        transaction.to_asset_id = to_asset_id if to_asset_id else None
        transaction.to_asset_rate = to_asset_rate
        transaction.commission_rate = commission_rate
        transaction.commission_type = commission_type
        transaction.save()
        
        return redirect('transactions')
    
    return render(request, 'transaction_form.html', {
        'transaction': transaction,
        'assets': assets,
        'transaction_types': [t for t in TransactionType.choices if t[0] != TransactionType.CHANGING_BALANCE],
        'transaction_categories': WasteCategory.choices,
        'refill_categories': RefillCategory.choices,
        'initial_date': transaction.date.strftime('%Y-%m-%d'),
        'initial_time': transaction.date.strftime('%H:%M'),
    })


@login_required
def assets(request):
    assets_list = Asset.objects.select_subclasses().filter(user=request.user, is_active=True)
    
    grouped = {}
    total_by_currency = {}
    
    for asset in assets_list:
        asset_type = get_asset_type_label(asset.type)
        if asset_type not in grouped:
            grouped[asset_type] = {'assets': [], 'totals_by_currency': {}}
        grouped[asset_type]['assets'].append(asset)
        
        if asset.currency not in grouped[asset_type]['totals_by_currency']:
            grouped[asset_type]['totals_by_currency'][asset.currency] = Decimal('0')
        grouped[asset_type]['totals_by_currency'][asset.currency] += asset.balance
        
        if asset.currency not in total_by_currency:
            total_by_currency[asset.currency] = Decimal('0')
        total_by_currency[asset.currency] += asset.balance
    
    return render(request, 'assets.html', {
        'assets_by_type': grouped,
        'total_by_currency': total_by_currency,
    })


@login_required
def asset_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        asset_type = request.POST.get('type')
        currency = request.POST.get('currency')
        balance = Decimal(request.POST.get('balance', '0'))

        if asset_type == AssetType.CASH:
            asset = CashAsset.objects.create(
                user=request.user,
                name=name,
                type=asset_type,
                currency=currency,
                location=request.POST.get('location', ''),
            )
        elif asset_type == AssetType.DEBIT_CARD:
            asset = DebitCardAsset.objects.create(
                user=request.user,
                name=name,
                type=asset_type,
                currency=currency,
                bank_id=request.POST.get('bank') or None,
                last_4_digits=request.POST.get('last_4_digits', ''),
            )
        elif asset_type == AssetType.CREDIT_CARD:
            asset = CreditCardAsset.objects.create(
                user=request.user,
                name=name,
                type=asset_type,
                currency=currency,
                bank_id=request.POST.get('bank') or None,
                credit_limit=request.POST.get('credit_limit') or None,
                grace_period_days=request.POST.get('grace_period_days') or None,
                last_4_digits=request.POST.get('last_4_digits', ''),
                billing_day=request.POST.get('billing_day') or None,
            )
        elif asset_type == AssetType.DEPOSIT:
            renewal_date = request.POST.get('renewal_date')
            asset = DepositAsset.objects.create(
                user=request.user,
                name=name,
                type=asset_type,
                currency=currency,
                bank_id=request.POST.get('bank') or None,
                interest_rate=request.POST.get('interest_rate') or None,
                term_months=request.POST.get('term_months') or None,
                renewal_date=renewal_date if renewal_date else None,
                is_capitalized=request.POST.get('is_capitalized') == 'on',
            )
        elif asset_type == AssetType.SAVING_ACCOUNT:
            asset = SavingAccount.objects.create(
                user=request.user,
                name=name,
                type=asset_type,
                currency=currency,
                bank_id=request.POST.get('bank') or None,
                interest_rate=request.POST.get('interest_rate') or None,
            )
        elif asset_type == AssetType.BROKERAGE:
            asset = BrokerageAsset.objects.create(
                user=request.user,
                name=name,
                type=asset_type,
                currency=currency,
                broker_name=request.POST.get('broker_name', ''),
                account_number=request.POST.get('account_number', ''),
                brokerage_account_type=request.POST.get('brokerage_account_type', ''),
            )
        elif asset_type == AssetType.E_WALLET:
            asset = EWalletAsset.objects.create(
                user=request.user,
                name=name,
                type=asset_type,
                currency=currency,
                provider_id=request.POST.get('provider') or None,
            )
        else:
            asset = Asset.objects.create(
                user=request.user,
                name=name,
                type=asset_type,
                currency=currency,
            )
        
        if balance != 0:
            Transaction.objects.create(
                user=request.user,
                type=TransactionType.CHANGING_BALANCE,
                amount=balance,
                currency=currency,
                to_asset=asset,
                to_asset_rate=Decimal('1'),
                description='Initial balance',
                date=timezone.now()
            )
        
        return redirect('assets')
    
    return render(request, 'asset_form.html', {
        'asset_types': AssetType.choices,
        'brokerage_account_types': BrokerageAccountType.choices,
        'banks': Bank.objects.all(),
        'providers': Provider.objects.all(),
    })


@login_required
def asset_edit(request, pk):
    asset = get_object_or_404(Asset, pk=pk, user=request.user)
    current_balance = asset.balance

    if asset.type == AssetType.CASH:
        asset = get_object_or_404(CashAsset, pk=pk)
    elif asset.type == AssetType.DEBIT_CARD:
        asset = get_object_or_404(DebitCardAsset, pk=pk)
    elif asset.type == AssetType.CREDIT_CARD:
        asset = get_object_or_404(CreditCardAsset, pk=pk)
    elif asset.type == AssetType.DEPOSIT:
        asset = get_object_or_404(DepositAsset, pk=pk)
    elif asset.type == AssetType.SAVING_ACCOUNT:
        asset = get_object_or_404(SavingAccount, pk=pk)
    elif asset.type == AssetType.BROKERAGE:
        asset = get_object_or_404(BrokerageAsset, pk=pk)
    elif asset.type == AssetType.E_WALLET:
        asset = get_object_or_404(EWalletAsset, pk=pk)
    
    if request.method == 'POST':
        asset.name = request.POST.get('name')
        asset.currency = request.POST.get('currency')
        asset.is_active = request.POST.get('is_active') == 'on'
        
        new_balance = Decimal(request.POST.get('balance', '0'))
        
        if asset.type == AssetType.CASH:
            asset.location = request.POST.get('location', '')
        elif asset.type == AssetType.DEBIT_CARD:
            asset.bank_id = request.POST.get('bank') or None
            asset.last_4_digits = request.POST.get('last_4_digits', '')
        elif asset.type == AssetType.CREDIT_CARD:
            asset.bank_id = request.POST.get('bank') or None
            asset.credit_limit = request.POST.get('credit_limit') or None
            asset.grace_period_days = request.POST.get('grace_period_days') or None
            asset.last_4_digits = request.POST.get('last_4_digits', '')
            asset.billing_day = request.POST.get('billing_day') or None
        elif asset.type == AssetType.DEPOSIT:
            asset.bank_id = request.POST.get('bank') or None
            asset.interest_rate = request.POST.get('interest_rate') or None
            asset.term_months = request.POST.get('term_months') or None
            asset.renewal_date = request.POST.get('renewal_date') or None
            asset.is_capitalized = request.POST.get('is_capitalized') == 'on'
        elif asset.type == AssetType.SAVING_ACCOUNT:
            asset.bank_id = request.POST.get('bank') or None
            asset.interest_rate = request.POST.get('interest_rate') or None
        elif asset.type == AssetType.BROKERAGE:
            asset.broker_name = request.POST.get('broker_name', '')
            asset.account_number = request.POST.get('account_number', '')
            asset.brokerage_account_type = request.POST.get('brokerage_account_type', '')
        elif asset.type == AssetType.E_WALLET:
            asset.provider_id = request.POST.get('provider') or None
        
        if new_balance != current_balance:
            balance_diff = new_balance - current_balance
            if balance_diff > 0:
                Transaction.objects.create(
                    user=request.user,
                    type=TransactionType.CHANGING_BALANCE,
                    amount=abs(balance_diff),
                    currency=asset.currency,
                    to_asset=asset,
                    to_asset_rate=Decimal('1'),
                    description='Balance correction',
                    date=timezone.now()
                )
            else:
                Transaction.objects.create(
                    user=request.user,
                    type=TransactionType.CHANGING_BALANCE,
                    amount=abs(balance_diff),
                    currency=asset.currency,
                    from_asset=asset,
                    from_asset_rate=Decimal('1'),
                    description='Balance correction',
                    date=timezone.now()
                )
        
        asset.save()
        return redirect('assets')
    
    return render(request, 'asset_form.html', {
        'asset': asset,
        'asset_types': AssetType.choices,
        'brokerage_account_types': BrokerageAccountType.choices,
        'banks': Bank.objects.all(),
        'providers': Provider.objects.all(),
    })


@login_required
def asset_delete(request, pk):
    asset = get_object_or_404(Asset, pk=pk, user=request.user)
    if request.method == 'POST':
        asset.delete()
        return redirect('assets')
    return HttpResponse(f"Delete asset: {asset.name}")


@login_required
def statistics(request, year=None, month=None, stat_type='outcome'):
    today = timezone.now()
    
    if stat_type not in ['income', 'outcome']:
        stat_type = 'outcome'
    
    if year is None:
        year = today.year
        month = today.month
        period = 'month'
    elif month is None:
        month = 1
        period = 'year'
    else:
        period = 'month'
    
    if period == 'year':
        start_date = make_aware(datetime(int(year), 1, 1))
        end_date = make_aware(datetime(int(year), 12, 31, 23, 59, 59))
        prev_year = int(year) - 1
        next_year = int(year) + 1
        prev_date = prev_year
        next_date = next_year
    else:
        current_date = make_aware(datetime(int(year), int(month), 1))
        start_date = current_date
        if int(month) == 12:
            end_date = make_aware(datetime(int(year) + 1, 1, 1)) - timedelta(seconds=1)
        else:
            from calendar import monthrange
            _, days = monthrange(int(year), int(month))
            end_date = make_aware(datetime(int(year), int(month), days, 23, 59, 59))
        prev_year = int(year)
        next_year = int(year)
        if int(month) == 1:
            prev_year = int(year) - 1
            prev_month = 12
        else:
            prev_month = int(month) - 1
        if int(month) == 12:
            next_year = int(year) + 1
            next_month = 1
        else:
            next_month = int(month) + 1
        prev_date = (prev_year, prev_month)
        next_date = (next_year, next_month)
    
    transactions_list = Transaction.objects.filter(
        user=request.user,
        date__gte=start_date,
        date__lte=end_date
    ).exclude(type=TransactionType.CHANGING_BALANCE)
    
    income_by_category = {}
    outcome_by_category = {}
    total_income = Decimal('0')
    total_outcome = Decimal('0')
    
    for t in transactions_list:
        if t.type == TransactionType.REFILL:
            cat = t.category if t.category else 'OTHER_REFILL'
            income_by_category[cat] = income_by_category.get(cat, Decimal('0')) + t.amount
            total_income += t.amount
        elif t.type == TransactionType.WASTE:
            cat = t.category if t.category else 'OTHER_WASTE'
            outcome_by_category[cat] = outcome_by_category.get(cat, Decimal('0')) + t.amount
            total_outcome += t.amount
    
    income_sorted = sorted(income_by_category.items(), key=lambda x: x[1], reverse=True)
    outcome_sorted = sorted(outcome_by_category.items(), key=lambda x: x[1], reverse=True)
    
    def format_category(cat):
        for choice in RefillCategory.choices:
            if choice[0] == cat:
                return choice[1]
        for choice in WasteCategory.choices:
            if choice[0] == cat:
                return choice[1]
        return cat
    
    income_data = []
    colors = ['#ff6b6b', '#ffa07a', '#ffd93d', '#6bcb77', '#4d96ff', '#9b59b6', '#ff9ff3', '#54a0ff', '#5f27cd', '#48dbfb', '#ff9f43', '#ee5a24', '#009432', '#1289A7', '#D980FA']
    for i, (cat, amount) in enumerate(income_sorted):
        percent = (amount / total_income * 100) if total_income > 0 else 0
        income_data.append({
            'category': format_category(cat),
            'amount': amount,
            'percent': percent,
            'color': colors[i % len(colors)],
        })
    
    outcome_data = []
    for i, (cat, amount) in enumerate(outcome_sorted):
        percent = (amount / total_outcome * 100) if total_outcome > 0 else 0
        outcome_data.append({
            'category': format_category(cat),
            'amount': amount,
            'percent': percent,
            'color': colors[i % len(colors)],
        })
    
    return render(request, 'statistics.html', {
        'year': int(year),
        'month': int(month),
        'period': period,
        'prev_year': prev_date[0] if period == 'month' else prev_date,
        'prev_month': prev_date[1] if period == 'month' else None,
        'next_year': next_date[0] if period == 'month' else next_date,
        'next_month': next_date[1] if period == 'month' else None,
        'stat_type': stat_type,
        'income_data': income_data,
        'outcome_data': outcome_data,
        'total_income': total_income,
        'total_outcome': total_outcome,
    })


@login_required
def profile(request):
    return render(request, 'profile.html', {
        'user': request.user,
    })


@login_required
def export_transactions(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    transactions = Transaction.objects.filter(
        user=request.user
    ).select_related('from_asset', 'to_asset').order_by('-date')

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ['Date', 'Type', 'Category', 'Amount', 'Currency', 'From Asset', 'To Asset', 'From Asset Rate', 'To Asset Rate', 'Commission Rate', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=t.date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=2, value=t.get_type_display())
        ws.cell(row=row, column=3, value=t.category or '')
        ws.cell(row=row, column=4, value=float(t.amount))
        ws.cell(row=row, column=5, value=t.currency)
        ws.cell(row=row, column=6, value=f'{t.from_asset.type}: {t.from_asset.name}' if t.from_asset else '')
        ws.cell(row=row, column=7, value=f'{t.to_asset.type}: {t.to_asset.name}' if t.to_asset else '')
        ws.cell(row=row, column=8, value=float(t.from_asset_rate) if t.from_asset_rate else 1.0)
        ws.cell(row=row, column=9, value=float(t.to_asset_rate) if t.to_asset_rate else 1.0)
        ws.cell(row=row, column=10, value=float(t.commission_rate) if t.commission_rate else 0.0)
        ws.cell(row=row, column=11, value=t.description or '')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
    wb.save(response)
    return response


import csv
import io


def import_transactions_csv(request, csv_file):
    assets = {f"{a.type}: {a.name}": a for a in Asset.objects.filter(user=request.user)}
    
    decoded = csv_file.read().decode('utf-8-sig')
    reader = csv.reader(io.StringIO(decoded))
    
    headers = next(reader)
    date_idx = headers.index('Date') if 'Date' in headers else 0
    type_idx = headers.index('Type') if 'Type' in headers else 1
    category_idx = headers.index('Category') if 'Category' in headers else 2
    amount_idx = headers.index('Amount') if 'Amount' in headers else 3
    currency_idx = headers.index('Currency') if 'Currency' in headers else 4
    from_asset_idx = headers.index('From Asset') if 'From Asset' in headers else 5
    to_asset_idx = headers.index('To Asset') if 'To Asset' in headers else 6
    from_asset_rate_idx = headers.index('From Asset Rate') if 'From Asset Rate' in headers else -1
    to_asset_rate_idx = headers.index('To Asset Rate') if 'To Asset Rate' in headers else -1
    commission_rate_idx = headers.index('Commission Rate') if 'Commission Rate' in headers else -1
    description_idx = headers.index('Description') if 'Description' in headers else 7
    
    imported_count = 0
    for row in reader:
        if not row[date_idx]:
            continue
        
        date_str = row[date_idx]
        if isinstance(date_str, str):
            t_date = make_aware(datetime.strptime(date_str, '%Y-%m-%d %H:%M'))
        else:
            t_date = make_aware(date_str)
        
        t_type_str = row[type_idx] if type_idx < len(row) else ''
        t_type = TransactionType.WASTE
        for choice in TransactionType.choices:
            if choice[1].lower() == t_type_str.lower():
                t_type = choice[0]
                break
        
        category = row[category_idx] if category_idx < len(row) and row[category_idx] else ''
        amount = Decimal(str(row[amount_idx])) if amount_idx < len(row) else Decimal('0')
        currency = row[currency_idx] if currency_idx < len(row) else 'RUB'
        
        from_asset_name = row[from_asset_idx] if from_asset_idx < len(row) else ''
        to_asset_name = row[to_asset_idx] if to_asset_idx < len(row) else ''
        description = row[description_idx] if description_idx < len(row) and row[description_idx] else ''
        
        from_asset = assets.get(from_asset_name) if from_asset_name else None
        to_asset = assets.get(to_asset_name) if to_asset_name else None
        
        if from_asset_rate_idx >= 0 and from_asset_rate_idx < len(row) and row[from_asset_rate_idx]:
            from_asset_rate = Decimal(str(row[from_asset_rate_idx]))
        else:
            from_asset_rate = Decimal('1')
        
        if to_asset_rate_idx >= 0 and to_asset_rate_idx < len(row) and row[to_asset_rate_idx]:
            to_asset_rate = Decimal(str(row[to_asset_rate_idx]))
        else:
            to_asset_rate = Decimal('1')
        
        if commission_rate_idx >= 0 and commission_rate_idx < len(row) and row[commission_rate_idx]:
            commission_rate = Decimal(str(row[commission_rate_idx]))
        else:
            commission_rate = Decimal('0')
        
        Transaction.objects.create(
            user=request.user,
            type=t_type,
            amount=amount,
            currency=currency,
            category=category,
            description=description,
            date=t_date,
            from_asset=from_asset,
            from_asset_rate=from_asset_rate,
            to_asset=to_asset,
            to_asset_rate=to_asset_rate,
            commission_rate=commission_rate,
        )
        imported_count += 1
    
    return render(request, 'profile.html', {
        'user': request.user,
        'success': f'Successfully imported {imported_count} transactions'
    })


@login_required
def import_transactions(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        if not excel_file:
            return render(request, 'profile.html', {
                'user': request.user,
                'error': 'No file selected'
            })
        
        try:
            file_name = excel_file.name.lower()
            if file_name.endswith('.csv'):
                return import_transactions_csv(request, excel_file)
            
            from openpyxl import load_workbook
            
            wb = load_workbook(excel_file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            date_idx = headers.index('Date') if 'Date' in headers else 0
            type_idx = headers.index('Type') if 'Type' in headers else 1
            category_idx = headers.index('Category') if 'Category' in headers else 2
            amount_idx = headers.index('Amount') if 'Amount' in headers else 3
            currency_idx = headers.index('Currency') if 'Currency' in headers else 4
            from_asset_idx = headers.index('From Asset') if 'From Asset' in headers else 5
            to_asset_idx = headers.index('To Asset') if 'To Asset' in headers else 6
            from_asset_rate_idx = headers.index('From Asset Rate') if 'From Asset Rate' in headers else -1
            to_asset_rate_idx = headers.index('To Asset Rate') if 'To Asset Rate' in headers else -1
            commission_rate_idx = headers.index('Commission Rate') if 'Commission Rate' in headers else -1
            description_idx = headers.index('Description') if 'Description' in headers else 7
            
            assets = {f"{a.type}: {a.name}": a for a in Asset.objects.filter(user=request.user)}
            
            imported_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[date_idx]:
                    continue
                
                date_str = row[date_idx]
                if isinstance(date_str, str):
                    t_date = make_aware(datetime.strptime(date_str, '%Y-%m-%d %H:%M'))
                else:
                    t_date = make_aware(date_str)
                
                t_type_str = row[type_idx] if type_idx < len(row) else ''
                t_type = TransactionType.WASTE
                for choice in TransactionType.choices:
                    if choice[1].lower() == t_type_str.lower():
                        t_type = choice[0]
                        break
                
                category = row[category_idx] if category_idx < len(row) and row[category_idx] else ''
                amount = Decimal(str(row[amount_idx])) if amount_idx < len(row) else Decimal('0')
                currency = row[currency_idx] if currency_idx < len(row) else 'RUB'
                
                from_asset_name = row[from_asset_idx] if from_asset_idx < len(row) else ''
                to_asset_name = row[to_asset_idx] if to_asset_idx < len(row) else ''
                description = row[description_idx] if description_idx < len(row) and row[description_idx] else ''
                
                from_asset = assets.get(from_asset_name) if from_asset_name else None
                to_asset = assets.get(to_asset_name) if to_asset_name else None
                
                if from_asset_rate_idx >= 0 and from_asset_rate_idx < len(row) and row[from_asset_rate_idx]:
                    from_asset_rate = Decimal(str(row[from_asset_rate_idx]))
                else:
                    from_asset_rate = Decimal('1')
                
                if to_asset_rate_idx >= 0 and to_asset_rate_idx < len(row) and row[to_asset_rate_idx]:
                    to_asset_rate = Decimal(str(row[to_asset_rate_idx]))
                else:
                    to_asset_rate = Decimal('1')
                
                if commission_rate_idx >= 0 and commission_rate_idx < len(row) and row[commission_rate_idx]:
                    commission_rate = Decimal(str(row[commission_rate_idx]))
                else:
                    commission_rate = Decimal('0')
                
                Transaction.objects.create(
                    user=request.user,
                    type=t_type,
                    amount=amount,
                    currency=currency,
                    category=category,
                    description=description,
                    date=t_date,
                    from_asset=from_asset,
                    from_asset_rate=from_asset_rate,
                    to_asset=to_asset,
                    to_asset_rate=to_asset_rate,
                    commission_rate=commission_rate,
                )
                imported_count += 1
            
            return render(request, 'profile.html', {
                'user': request.user,
                'success': f'Successfully imported {imported_count} transactions'
            })
        except Exception as e:
            return render(request, 'profile.html', {
                'user': request.user,
                'error': f'Error importing file: {str(e)}'
            })
    
    return redirect('profile')


@login_required
def transaction_delete(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, user=request.user)
    if request.method == 'POST':
        transaction.delete()
        return redirect('transactions')
    return HttpResponse(f"Delete transaction: {transaction.id}")


def api_exchange_rate(request):
    from_currency = request.GET.get('from', 'RUB')
    to_currency = request.GET.get('to', 'RUB')
    date_str = request.GET.get('date')

    at_date = None
    if date_str:
        try:
            at_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    rate = ExchangeRateService.get_rate(from_currency, to_currency, at_date)
    return JsonResponse({'rate': str(rate)})


@login_required
def banks(request):
    banks_list = Bank.objects.all()
    providers_list = Provider.objects.all()
    return render(request, 'banks.html', {'banks': banks_list, 'providers': providers_list})


@login_required
def bank_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        image = request.FILES.get('image')
        Bank.objects.create(name=name, image=image)
        return redirect('banks')
    return render(request, 'bank_form.html')


@login_required
def bank_edit(request, pk):
    bank = get_object_or_404(Bank, pk=pk)
    if request.method == 'POST':
        bank.name = request.POST.get('name')
        if request.FILES.get('image'):
            bank.image = request.FILES.get('image')
        bank.cashback_categories_limit = int(request.POST.get('cashback_categories_limit', 3))
        bank.save()
        return redirect('banks')
    return render(request, 'bank_form.html', {'bank': bank})


@login_required
def bank_view(request, pk):
    bank = get_object_or_404(Bank, pk=pk)
    assets = []
    for asset_class in [DebitCardAsset, CreditCardAsset, DepositAsset, SavingAccount]:
        assets.extend(asset_class.objects.filter(bank=bank, user=request.user))

    today = timezone.now()
    current_year = today.year
    current_month = today.month

    if current_month == 12:
        next_year = current_year + 1
        next_month = 1
    else:
        next_year = current_year
        next_month = current_month + 1

    current_month_obj = BankCashbackMonth.objects.filter(bank=bank, year=current_year, month=current_month).first()
    next_month_obj = BankCashbackMonth.objects.filter(bank=bank, year=next_year, month=next_month).first()

    def get_month_data(month_obj):
        """Get selected and available categories for a specific month."""
        if not month_obj:
            return None, [], [], {}, {}
        
        # Get selections for this month
        selections = BankCashbackSelection.objects.filter(
            bank_cashback_month=month_obj,
            is_selected=True
        ).select_related('bank_cashback_category__category')
        
        selected_ids = set(s.bank_cashback_category_id for s in selections)
        
        # Selected bank cashback categories
        selected = [s.bank_cashback_category for s in selections]
        
        # Get month-specific category configs
        month_cat_configs = BankCashbackMonthCategory.objects.filter(
            bank_cashback_month=month_obj
        )
        month_category_ids = {}
        month_category_values = {}
        for mc in month_cat_configs:
            month_category_ids[mc.category_id] = True
            month_category_values[mc.category_id] = {
                'percent': mc.percent,
                'limit': mc.limit
            }
        
        # Available: BankCashbackCategory objects for this bank that are not selected
        available = BankCashbackCategory.objects.filter(
            bank=bank
        ).exclude(id__in=selected_ids).select_related('category')
        
        return month_obj, selected, list(available), month_category_ids, month_category_values

    current_month_obj, current_selected, current_available, current_month_cat_ids, current_month_cat_values = get_month_data(current_month_obj)
    next_month_obj, next_selected, next_available, next_month_cat_ids, next_month_cat_values = get_month_data(next_month_obj)

    # Get all cashback categories for the modal
    all_cashback_categories = CashbackCategory.objects.all()

    return render(request, 'bank_view.html', {
        'bank': bank,
        'assets': assets,
        'current_year': current_year,
        'current_month': current_month,
        'next_year': next_year,
        'next_month': next_month,
        'current_month_obj': current_month_obj,
        'next_month_obj': next_month_obj,
        'current_selected': current_selected,
        'current_available': current_available,
        'next_selected': next_selected,
        'next_available': next_available,
        'current_month_category_ids': current_month_cat_ids,
        'next_month_category_ids': next_month_cat_ids,
        'current_month_category_values': current_month_cat_values,
        'next_month_category_values': next_month_cat_values,
        'all_cashback_categories': all_cashback_categories,
    })


@login_required
def add_cashback_categories(request, pk, year, month):
    bank = get_object_or_404(Bank, pk=pk)

    cashback_month, created = BankCashbackMonth.objects.get_or_create(
        bank=bank,
        year=int(year),
        month=int(month),
        defaults={'common_limit': None, 'max_categories': None}
    )

    # Get all cashback categories and create bank-specific ones
    for cashback_cat in CashbackCategory.objects.all():
        BankCashbackCategory.objects.get_or_create(
            bank=bank,
            category=cashback_cat,
            defaults={
                'percent': Decimal('1.0'),
                'limit': None
            }
        )

    return redirect('bank_view', pk=pk)


@login_required
def select_cashback_category(request, pk, year, month, category_id):
    bank = get_object_or_404(Bank, pk=pk)
    cashback_month = get_object_or_404(BankCashbackMonth, bank=bank, year=int(year), month=int(month))
    category = get_object_or_404(BankCashbackCategory, pk=category_id, bank=bank)

    selection, created = BankCashbackSelection.objects.get_or_create(
        bank_cashback_month=cashback_month,
        bank_cashback_category=category,
        defaults={'is_selected': True}
    )

    selected_count = BankCashbackSelection.objects.filter(
        bank_cashback_month=cashback_month,
        is_selected=True
    ).count()

    max_categories = cashback_month.get_max_categories()

    if created:
        if selected_count > max_categories:
            selection.is_selected = False
            selection.save()
        return redirect('bank_view', pk=pk)

    if selection.is_selected:
        selection.is_selected = False
        selection.save()
    elif selected_count < max_categories:
        selection.is_selected = True
        selection.save()

    return redirect('bank_view', pk=pk)


@login_required
def bank_cashback_edit(request, pk, year, month):
    bank = get_object_or_404(Bank, pk=pk)
    cashback_month = get_object_or_404(BankCashbackMonth, bank=bank, year=int(year), month=int(month))

    if request.method == 'POST':
        common_limit = request.POST.get('common_limit')
        cashback_month.common_limit = Decimal(common_limit) if common_limit else None
        max_categories = request.POST.get('max_categories')
        cashback_month.max_categories = int(max_categories) if max_categories else None
        cashback_month.save()
        return redirect('bank_view', pk=pk)

    return render(request, 'bank_cashback_edit.html', {
        'bank': bank,
        'cashback_month': cashback_month,
    })


@login_required
def provider_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        image = request.FILES.get('image')
        Provider.objects.create(name=name, image=image)
        return redirect('banks')
    return render(request, 'provider_form.html')


@login_required
def provider_edit(request, pk):
    provider = get_object_or_404(Provider, pk=pk)
    if request.method == 'POST':
        provider.name = request.POST.get('name')
        if request.FILES.get('image'):
            provider.image = request.FILES.get('image')
        provider.save()
        return redirect('banks')
    return render(request, 'provider_form.html', {'provider': provider})


@login_required
def provider_view(request, pk):
    provider = get_object_or_404(Provider, pk=pk)
    assets = EWalletAsset.objects.filter(provider=provider, user=request.user)
    return render(request, 'provider_view.html', {'provider': provider, 'assets': assets})


@login_required
def cashback_overview(request, year=None, month=None):
    today = timezone.now()
    if year is None:
        year = today.year
    if month is None:
        month = today.month

    year = int(year)
    month = int(month)

    if month == 1:
        prev_year = year - 1
        prev_month = 12
    else:
        prev_year = year
        prev_month = month - 1

    if month == 12:
        next_year = year + 1
        next_month = 1
    else:
        next_year = year
        next_month = month + 1

    banks_with_cashback = []
    all_banks = Bank.objects.all()

    for bank in all_banks:
        cashback_month = BankCashbackMonth.objects.filter(bank=bank, year=year, month=month).first()
        if cashback_month:
            selections = BankCashbackSelection.objects.filter(
                bank_cashback_month=cashback_month,
                is_selected=True
            ).select_related('bank_cashback_category__category')
            
            selected_ids = set(s.bank_cashback_category_id for s in selections)
            categories = [s.bank_cashback_category for s in selections]
            
            # Get all BankCashbackCategory for this bank, excluding selected ones
            all_bank_cats = BankCashbackCategory.objects.filter(
                bank=bank
            ).exclude(id__in=selected_ids).select_related('category')
            
            banks_with_cashback.append({
                'bank': bank,
                'cashback_month': cashback_month,
                'categories': categories,
                'selected_count': len(categories),
                'max_categories': cashback_month.get_max_categories(),
                'available': list(all_bank_cats),
                'selected_ids': selected_ids,
            })

    return render(request, 'cashback_overview.html', {
        'year': year,
        'month': month,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
        'banks_with_cashback': banks_with_cashback,
    })


@login_required
def cashback_overview_save(request, pk, year, month):
    """
    Save category selections from cashback overview page.
    """
    bank = get_object_or_404(Bank, pk=pk)
    cashback_month = get_object_or_404(BankCashbackMonth, bank=bank, year=int(year), month=int(month))
    
    if request.method == 'POST':
        selected_ids_str = request.POST.getlist('selected_ids[]')
        selected_ids = [int(cid) for cid in selected_ids_str if cid.strip()]
        
        # Deselect all first
        BankCashbackSelection.objects.filter(
            bank_cashback_month=cashback_month
        ).update(is_selected=False)
        
        # Select the chosen ones (up to max limit)
        max_categories = cashback_month.get_max_categories()
        selected_count = 0
        
        for bank_cat_id in selected_ids:
            if selected_count >= max_categories:
                break
            try:
                bank_cat = BankCashbackCategory.objects.get(pk=bank_cat_id, bank=bank)
                BankCashbackSelection.objects.update_or_create(
                    bank_cashback_month=cashback_month,
                    bank_cashback_category=bank_cat,
                    defaults={'is_selected': True}
                )
                selected_count += 1
            except BankCashbackCategory.DoesNotExist:
                pass
    
    return redirect('cashback_overview_month', year=year, month=month)


@login_required
def cashback_overview_select(request, year, month, bank_id):
    bank = get_object_or_404(Bank, pk=bank_id)
    cashback_month = get_object_or_404(BankCashbackMonth, bank=bank, year=int(year), month=int(month))
    category_id = request.POST.get('category_id')

    if category_id:
        category = get_object_or_404(BankCashbackCategory, pk=category_id, bank=bank)
        selection, created = BankCashbackSelection.objects.get_or_create(
            bank_cashback_month=cashback_month,
            bank_cashback_category=category,
            defaults={'is_selected': True}
        )

        selected_count = BankCashbackSelection.objects.filter(
            bank_cashback_month=cashback_month,
            is_selected=True
        ).count()

        max_categories = cashback_month.get_max_categories()

        if created:
            if selected_count > max_categories:
                selection.is_selected = False
                selection.save()
        elif selection.is_selected:
            selection.is_selected = False
            selection.save()
        elif selected_count < max_categories:
            selection.is_selected = True
            selection.save()

    return redirect('cashback_overview', year=year, month=month)


@login_required
def bank_save_categories(request, pk, year, month):
    """
    Step 1: Save selected categories with their percents and limits for a specific month.
    This creates/updates BankCashbackMonthCategory records and redirects back to bank view.
    """
    bank = get_object_or_404(Bank, pk=pk)
    
    # Get or create BankCashbackMonth for this month
    cashback_month, created = BankCashbackMonth.objects.get_or_create(
        bank=bank,
        year=int(year),
        month=int(month),
        defaults={'common_limit': None, 'max_categories': None}
    )

    if request.method == 'POST':
        # Get the list of category IDs from the form
        category_ids = request.POST.getlist('category_ids[]')
        
        # First, remove categories not in the list
        BankCashbackMonthCategory.objects.filter(
            bank_cashback_month=cashback_month
        ).exclude(category_id__in=category_ids).delete()
        
        # Then create/update the selected categories
        for cat_id in category_ids:
            category = get_object_or_404(CashbackCategory, pk=cat_id)
            percent = Decimal(request.POST.get(f'percent_{cat_id}', '0'))
            limit = request.POST.get(f'limit_{cat_id}')
            limit = Decimal(limit) if limit else None
            
            # Create/update month-specific category config
            BankCashbackMonthCategory.objects.update_or_create(
                bank_cashback_month=cashback_month,
                category=category,
                defaults={
                    'percent': percent,
                    'limit': limit,
                }
            )
            
            # Also create/update bank-level category (for use in selections)
            BankCashbackCategory.objects.update_or_create(
                bank=bank,
                category=category,
                defaults={
                    'percent': percent,
                    'limit': limit,
                }
            )
    
    return redirect('bank_view', pk=pk)


@login_required
def bank_add_new_category(request, pk):
    """
    Add a new cashback category globally.
    """
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            # Just create the CashbackCategory - it will be available for selection
            CashbackCategory.objects.get_or_create(name=name)
    
    return redirect(request.POST.get('next', 'bank_view'), pk=pk)


@login_required
def bank_select_categories(request, pk, year, month):
    """
    Step 2: Select categories for the month (up to max_categories limit).
    """
    bank = get_object_or_404(Bank, pk=pk)
    
    cashback_month, created = BankCashbackMonth.objects.get_or_create(
        bank=bank,
        year=int(year),
        month=int(month),
        defaults={'common_limit': None, 'max_categories': None}
    )
    
    # Get all categories available for this bank
    bank_categories = BankCashbackCategory.objects.filter(bank=bank).select_related('category')
    
    # Get currently selected categories for this month
    selections = BankCashbackSelection.objects.filter(
        bank_cashback_month=cashback_month,
        is_selected=True
    ).select_related('bank_cashback_category')
    selected_ids = set(s.bank_cashback_category_id for s in selections)
    
    if request.method == 'POST':
        # Handle selection/unselection
        selected_category_ids = set(int(cid) for cid in request.POST.getlist('selected_categories[]'))
        
        max_categories = cashback_month.get_max_categories()
        
        # First, deselect all current selections
        BankCashbackSelection.objects.filter(
            bank_cashback_month=cashback_month
        ).update(is_selected=False)
        
        # Then select up to max_categories
        selected_count = 0
        for cat in bank_categories:
            if cat.id in selected_category_ids and selected_count < max_categories:
                BankCashbackSelection.objects.update_or_create(
                    bank_cashback_month=cashback_month,
                    bank_cashback_category=cat,
                    defaults={'is_selected': True}
                )
                selected_count += 1
        
        return redirect('bank_view', pk=pk)
    
    # Prepare data for template
    categories_with_selection = []
    for cat in bank_categories:
        categories_with_selection.append({
            'bank_category': cat,
            'is_selected': cat.id in selected_ids,
        })
    
    return render(request, 'bank_select_categories.html', {
        'bank': bank,
        'cashback_month': cashback_month,
        'year': year,
        'month': month,
        'categories': categories_with_selection,
        'max_categories': cashback_month.get_max_categories(),
        'selected_count': len(selected_ids),
    })


@login_required
def bank_save_month_selection(request, pk, year, month):
    """
    Save the month's category selections.
    """
    bank = get_object_or_404(Bank, pk=pk)
    
    cashback_month, created = BankCashbackMonth.objects.get_or_create(
        bank=bank,
        year=int(year),
        month=int(month),
        defaults={'common_limit': None, 'max_categories': None}
    )
    
    if request.method == 'POST':
        selected_ids_str = request.POST.get('selected_ids', '')
        selected_ids = [int(cid) for cid in selected_ids_str.split(',') if cid.strip()]
        
        # Deselect all first
        BankCashbackSelection.objects.filter(
            bank_cashback_month=cashback_month
        ).update(is_selected=False)
        
        # Select the chosen ones (up to max limit)
        max_categories = cashback_month.get_max_categories()
        selected_count = 0
        
        for bank_cat_id in selected_ids:
            if selected_count >= max_categories:
                break
            try:
                bank_cat = BankCashbackCategory.objects.get(pk=bank_cat_id, bank=bank)
                BankCashbackSelection.objects.update_or_create(
                    bank_cashback_month=cashback_month,
                    bank_cashback_category=bank_cat,
                    defaults={'is_selected': True}
                )
                selected_count += 1
            except BankCashbackCategory.DoesNotExist:
                pass
    
    return redirect('bank_view', pk=pk)


@login_required
def cashback_categories_list(request):
    categories = CashbackCategory.objects.all()
    return render(request, 'cashback_categories_list.html', {
        'categories': categories,
    })


@login_required
def cashback_category_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        icon = request.POST.get('icon', '🛒')
        color = request.POST.get('color', '#4CAF50')
        
        if not name:
            return redirect('cashback_category_create')
        
        if CashbackCategory.objects.filter(name=name).exists():
            return redirect('cashback_category_create')
        
        svg_content = f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64" fill="none">
  <rect x="8" y="8" width="48" height="48" rx="8" fill="{color}"/>
  <text x="32" y="40" font-family="Arial" font-size="24" fill="white" text-anchor="middle">{icon}</text>
</svg>'''
        
        safe_name = name.lower().replace(' ', '_').replace("'", '')
        filename = f'cashback_categories/{safe_name}.svg'
        filepath = os.path.join(settings.STATIC_ROOT, filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'w') as f:
            f.write(svg_content)
        
        static_dir = settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else os.path.join(settings.BASE_DIR, 'static')
        shutil.copy(filepath, os.path.join(static_dir, filename))
        
        category = CashbackCategory.objects.create(
            name=name,
            image=filename
        )
        
        return redirect('cashback_categories_list')
    
    icons = [
        '🛒', '🍽️', '🍔', '💊', '🚕', '🛍️', '❤️', '⛽', '🏠', '🔧',
        '🐾', '🐕', '🚖', '⚽', '🎭', '📚', '✈️', '🌍', '👶', '💐',
        '🎨', '📖', '💪', '📱', '🧹', '📦', '🏪', '🎯', '🎲', '🎁',
        '🍕', '☕', '🍺', '🍷', '🎮', '🎸', '🎬', '📷', '🏋️', '🧘',
        '💇', '💅', '👗', '👠', '⌚', '💻', '📺', '🎧', '🎤', '🏆',
        '🥇', '🎪', '🎡', '🎢', '🏕️', '🏖️', '🗽', '🏰', '⛺', '🌴',
        '🍜', '🍣', '🍱', '🥗', '🍰', '🧁', '🍩', '🍪', '🥤', '🍵',
        '🌸', '🌺', '🌻', '🌹', '🥀', '🌷', '🌱', '🌲', '🍎', '🍇',
        '🍕', '🌭', '🍟', '🌮', '🌯', '🥙', '🥪', '🍞', '🥐', '🥯',
        '💰', '💳', '💵', '💎', '🪙', '💴', '💶', '💷', '🏦', '🏧',
        '🛠️', '🔩', '⚙️', '🔧', '🔨', '⛏️', '🔪', '🗡️', '⚔️', '🔫',
        '💡', '🔮', '🧲', '🔭', '🔬', '🧪', '🧬', '🩺', '🩻', '🩭',
        '🎓', '📝', '📒', '📓', '📔', '📕', '📗', '📘', '📙', '📚',
        '🎒', '👝', '👛', '👜', '🎒', '🧳', '👛', '💼', '📁', '📂',
    ]
    colors = [
        '#4CAF50', '#E91E63', '#FF9800', '#2196F3', '#9C27B0', '#F44336',
        '#607D8B', '#795548', '#9E9E9E', '#FF5722', '#8D6E63', '#673AB7',
        '#3F51B5', '#03A9F4', '#00BCD4', '#CDDC39', '#FFEB3B', '#FFC107',
        '#FF9800', '#FF5722', '#E91E63', '#9C27B0', '#673AB7', '#3F51B5',
        '#2196F3', '#03A9F4', '#00BCD4', '#009688', '#4CAF50', '#8BC34A',
        '#CDDC39', '#FFC107', '#FF9800', '#FF5722', '#795548', '#9E9E9E',
        '#607D8B', '#263238', '#ECEFF1', '#CFD8DC', '#B0BEC5', '#90A4AE',
        '#78909C', '#607D8B', '#546E7A', '#455A64', '#37474F', '#263238',
        '#FFCDD2', '#F8BBD9', '#E1BEE7', '#D1C4E9', '#C5CAE9', '#BBDEFB',
        '#B3E5FC', '#B2EBF2', '#B2DFDB', '#C8E6C9', '#DCEDC8', '#F0F4C3',
        '#FFF9C4', '#FFECB3', '#FFE0B2', '#FFCCBC', '#D7CCC8', '#CFD8DC',
    ]
    
    return render(request, 'cashback_category_create.html', {
        'icons': icons,
        'colors': colors,
    })


@login_required
def cashback_category_edit(request, pk):
    category = get_object_or_404(CashbackCategory, pk=pk)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        icon = request.POST.get('icon', '🛒')
        color = request.POST.get('color', '#4CAF50')
        
        if name:
            category.name = name
        
        svg_content = f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64" fill="none">
  <rect x="8" y="8" width="48" height="48" rx="8" fill="{color}"/>
  <text x="32" y="40" font-family="Arial" font-size="24" fill="white" text-anchor="middle">{icon}</text>
</svg>'''
        
        filename = f'cashback_categories/{category.name.lower().replace(" ", "_")}.svg'
        filepath = os.path.join(settings.STATIC_ROOT, filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'w') as f:
            f.write(svg_content)
        
        shutil.copy(filepath, os.path.join(settings.STATICFILES_DIRS[0], filename))
        
        category.image = filename
        category.save()
        
        return redirect('cashback_categories_list')
    
    icons = [
        '🛒', '🍽️', '🍔', '💊', '🚕', '🛍️', '❤️', '⛽', '🏠', '🔧',
        '🐾', '🐕', '🚖', '⚽', '🎭', '📚', '✈️', '🌍', '👶', '💐',
        '🎨', '📖', '💪', '📱', '🧹', '📦', '🏪', '🎯', '🎲', '🎁',
        '🍕', '☕', '🍺', '🍷', '🎮', '🎸', '🎬', '📷', '🏋️', '🧘',
        '💇', '💅', '👗', '👠', '⌚', '💻', '📺', '🎧', '🎤', '🏆',
        '🥇', '🎪', '🎡', '🎢', '🏕️', '🏖️', '🗽', '🏰', '⛺', '🌴',
        '🍜', '🍣', '🍱', '🥗', '🍰', '🧁', '🍩', '🍪', '🥤', '🍵',
        '🌸', '🌺', '🌻', '🌹', '🥀', '🌷', '🌱', '🌲', '🍎', '🍇',
        '🍕', '🌭', '🍟', '🌮', '🌯', '🥙', '🥪', '🍞', '🥐', '🥯',
        '💰', '💳', '💵', '💎', '🪙', '💴', '💶', '💷', '🏦', '🏧',
        '🛠️', '🔩', '⚙️', '🔧', '🔨', '⛏️', '🔪', '🗡️', '⚔️', '🔫',
        '💡', '🔮', '🧲', '🔭', '🔬', '🧪', '🧬', '🩺', '🩻', '🩭',
        '🎓', '📝', '📒', '📓', '📔', '📕', '📗', '📘', '📙', '📚',
        '🎒', '👝', '👛', '👜', '🎒', '🧳', '👛', '💼', '📁', '📂',
    ]
    colors = [
        '#4CAF50', '#E91E63', '#FF9800', '#2196F3', '#9C27B0', '#F44336',
        '#607D8B', '#795548', '#9E9E9E', '#FF5722', '#8D6E63', '#673AB7',
        '#3F51B5', '#03A9F4', '#00BCD4', '#CDDC39', '#FFEB3B', '#FFC107',
        '#FF9800', '#FF5722', '#E91E63', '#9C27B0', '#673AB7', '#3F51B5',
        '#2196F3', '#03A9F4', '#00BCD4', '#009688', '#4CAF50', '#8BC34A',
        '#CDDC39', '#FFC107', '#FF9800', '#FF5722', '#795548', '#9E9E9E',
        '#607D8B', '#263238', '#ECEFF1', '#CFD8DC', '#B0BEC5', '#90A4AE',
        '#78909C', '#607D8B', '#546E7A', '#455A64', '#37474F', '#263238',
        '#FFCDD2', '#F8BBD9', '#E1BEE7', '#D1C4E9', '#C5CAE9', '#BBDEFB',
        '#B3E5FC', '#B2EBF2', '#B2DFDB', '#C8E6C9', '#DCEDC8', '#F0F4C3',
        '#FFF9C4', '#FFECB3', '#FFE0B2', '#FFCCBC', '#D7CCC8', '#CFD8DC',
    ]
    
    return render(request, 'cashback_category_edit.html', {
        'category': category,
        'icons': icons,
        'colors': colors,
    })
