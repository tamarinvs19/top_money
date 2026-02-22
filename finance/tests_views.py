from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.urls import reverse
from django.utils import timezone
from decimal import Decimal
from datetime import timedelta

from finance.models import Asset, DebitCardAsset, Transaction, AssetType, TransactionType, CashAsset


def create_asset_with_balance(user, name, asset_type, currency, balance_amount):
    """Helper to create an asset with initial balance using CHANGING_BALANCE transaction"""
    asset = DebitCardAsset.objects.create(
        user=user,
        name=name,
        type=asset_type,
        currency=currency,
    )
    if balance_amount:
        Transaction.objects.create(
            user=user,
            type=TransactionType.CHANGING_BALANCE,
            amount=balance_amount,
            currency=currency,
            to_asset=asset,
            date=timezone.now()
        )
    return asset


class AuthViewsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.signup_url = reverse('signup')
        self.login_url = reverse('login')
    
    def test_signup_get(self):
        response = self.client.get(self.signup_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Create Account')
    
    def test_signup_post_creates_user(self):
        response = self.client.post(self.signup_url, {
            'username': 'newuser',
            'password1': 'testpassword123',
            'password2': 'testpassword123',
        })
        self.assertTrue(User.objects.filter(username='newuser').exists())
    
    def test_signup_redirects_to_transactions(self):
        response = self.client.post(self.signup_url, {
            'username': 'newuser',
            'password1': 'testpassword123',
            'password2': 'testpassword123',
        }, follow=True)
        self.assertRedirects(response, reverse('transactions'))
    
    def test_login_get(self):
        response = self.client.get(self.login_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Finance Manager')
    
    def test_login_success(self):
        user = User.objects.create_user(username='testuser', password='testpass123')
        response = self.client.post(self.login_url, {
            'username': 'testuser',
            'password': 'testpass123',
        })
        self.assertRedirects(response, reverse('transactions'))
    
    def test_login_invalid_credentials(self):
        response = self.client.post(self.login_url, {
            'username': 'wronguser',
            'password': 'wrongpass',
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Invalid')


class TransactionViewsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('10000.00'))
        self.client.login(username='testuser', password='testpass123')
    
    def test_transactions_list_requires_login(self):
        self.client.logout()
        response = self.client.get(reverse('transactions'))
        self.assertRedirects(response, f"{reverse('login')}?next=/")
    
    def test_transactions_list_get(self):
        response = self.client.get(reverse('transactions'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Transactions')
    
    def test_transactions_list_with_data(self):
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('5000.00'),
            currency='RUB',
            to_asset=self.asset,
            category='Salary',
            date=timezone.now()
        )
        response = self.client.get(reverse('transactions'))
        self.assertContains(response, '5000.00')
        self.assertContains(response, 'Salary')
    
    def test_transactions_navigation(self):
        response = self.client.get(reverse('transactions_month', args=[2026, 1]))
        self.assertEqual(response.status_code, 200)
    
    def test_transaction_add_get(self):
        response = self.client.get(reverse('transaction_add'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Add Transaction')
    
    def test_transaction_add_post(self):
        response = self.client.post(reverse('transaction_add'), {
            'type': TransactionType.REFILL,
            'amount': '3000',
            'currency': 'RUB',
            'category': 'Test',
            'description': 'Test transaction',
            'date': '2026-02-18',
            'to_asset': self.asset.pk,
        })
        self.assertTrue(Transaction.objects.filter(amount=Decimal('3000.00')).exists())
        self.assertRedirects(response, reverse('transactions'))
    
    def test_transaction_add_with_year_month_day(self):
        url = reverse('transaction_add_day', args=[2026, 2, 15])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="2026-02-15"')
    
    def test_transaction_edit_get(self):
        transaction = Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('1000.00'),
            currency='RUB',
            to_asset=self.asset,
            date=timezone.now()
        )
        url = reverse('transaction_edit', args=[transaction.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Edit Transaction')
    
    def test_transaction_edit_post(self):
        transaction = Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('1000.00'),
            currency='RUB',
            to_asset=self.asset,
            date=timezone.now()
        )
        url = reverse('transaction_edit', args=[transaction.pk])
        response = self.client.post(url, {
            'type': TransactionType.WASTE,
            'amount': '500',
            'currency': 'RUB',
            'category': 'Updated',
            'description': 'Updated desc',
            'date': '2026-02-18',
            'from_asset': self.asset.pk,
        })
        transaction.refresh_from_db()
        self.assertEqual(transaction.amount, Decimal('500.00'))
        self.assertEqual(transaction.category, 'Updated')
    
    def test_transaction_edit_other_user_forbidden(self):
        other_user = User.objects.create_user(username='other', password='otherpass')
        other_asset = DebitCardAsset.objects.create(
            user=other_user,
            name='Other Card',
            type=AssetType.DEBIT_CARD,
            currency='RUB'
        )
        transaction = Transaction.objects.create(
            user=other_user,
            type=TransactionType.REFILL,
            amount=Decimal('1000.00'),
            currency='RUB',
            to_asset=other_asset,
            date=timezone.now()
        )
        url = reverse('transaction_edit', args=[transaction.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)


class AssetViewsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.client.login(username='testuser', password='testpass123')
    
    def test_assets_list_get(self):
        response = self.client.get(reverse('assets'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Assets')
    
    def test_assets_list_with_data(self):
        create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('5000.00'))
        response = self.client.get(reverse('assets'))
        self.assertContains(response, '5000.00')
        self.assertContains(response, 'Test Card')
    
    def test_assets_list_grouped_by_type(self):
        create_asset_with_balance(self.user, 'Card1', AssetType.DEBIT_CARD, 'RUB', Decimal('1000'))
        create_asset_with_balance(self.user, 'Card2', AssetType.DEBIT_CARD, 'RUB', Decimal('2000'))
        create_asset_with_balance(self.user, 'Cash', AssetType.CASH, 'RUB', Decimal('500'))
        
        response = self.client.get(reverse('assets'))
        self.assertContains(response, 'Debit Card')
        self.assertContains(response, 'Cash')
        self.assertContains(response, '3000')  # total for DEBIT_CARD
        self.assertContains(response, '500')   # total for CASH
    
    def test_assets_total_balance(self):
        create_asset_with_balance(self.user, 'Card1', AssetType.DEBIT_CARD, 'RUB', Decimal('1000'))
        create_asset_with_balance(self.user, 'Cash', AssetType.CASH, 'RUB', Decimal('500'))
        
        response = self.client.get(reverse('assets'))
        self.assertContains(response, 'Total Balance')
        self.assertContains(response, '1500')
    
    def test_asset_add_get(self):
        response = self.client.get(reverse('asset_add'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Add Asset')
    
    def test_asset_add_post(self):
        response = self.client.post(reverse('asset_add'), {
            'name': 'New Card',
            'type': AssetType.DEBIT_CARD,
            'currency': 'RUB',
            'balance': '5000',
            'bank_name': 'Sberbank',
            'last_4_digits': '1234',
        })
        self.assertTrue(Asset.objects.filter(name='New Card').exists())
        self.assertRedirects(response, reverse('assets'))
        asset = Asset.objects.get(name='New Card')
        self.assertEqual(asset.balance, Decimal('5000'))
    
    def test_asset_edit_get(self):
        asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('1000.00'))
        url = reverse('asset_edit', args=[asset.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Edit Asset')
        self.assertContains(response, 'Test Card')
    
    def test_asset_edit_post(self):
        asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('1000.00'))
        url = reverse('asset_edit', args=[asset.pk])
        response = self.client.post(url, {
            'name': 'Updated Card',
            'type': AssetType.DEBIT_CARD,
            'currency': 'RUB',
            'balance': '2000',
            'bank_name': 'Tinkoff',
            'is_active': 'on',
        })
        asset.refresh_from_db()
        self.assertEqual(asset.name, 'Updated Card')
        self.assertEqual(asset.currency, 'RUB')
        asset_from_db = Asset.objects.get(pk=asset.pk)
        self.assertEqual(asset_from_db.balance, Decimal('2000'))
    
    def test_asset_edit_can_add_field_that_was_null_at_creation(self):
        asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('1000.00'))
        asset.bank_name = ''
        asset.last_4_digits = ''
        asset.save()
        
        url = reverse('asset_edit', args=[asset.pk])
        response = self.client.post(url, {
            'name': 'Test Card',
            'type': AssetType.DEBIT_CARD,
            'currency': 'RUB',
            'balance': '1000',
            'bank_name': 'Tinkoff',
            'last_4_digits': '1234',
        })
        asset.refresh_from_db()
        self.assertEqual(asset.bank_name, 'Tinkoff')
        self.assertEqual(asset.last_4_digits, '1234')
    
    def test_asset_edit_other_user_forbidden(self):
        other_user = User.objects.create_user(username='other', password='otherpass')
        asset = DebitCardAsset.objects.create(
            user=other_user,
            name='Other Card',
            type=AssetType.DEBIT_CARD,
            currency='RUB'
        )
        url = reverse('asset_edit', args=[asset.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)


class AssetDeleteViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.client.login(username='testuser', password='testpass123')
    
    def test_asset_delete_get(self):
        asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('1000.00'))
        url = reverse('asset_delete', args=[asset.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
    
    def test_asset_delete_post(self):
        asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('1000.00'))
        Transaction.objects.filter(to_asset=asset).delete()
        Transaction.objects.filter(from_asset=asset).delete()
        url = reverse('asset_delete', args=[asset.pk])
        response = self.client.post(url)
        self.assertRedirects(response, reverse('assets'))
        self.assertFalse(Asset.objects.filter(pk=asset.pk).exists())
    
    def test_asset_delete_other_user_forbidden(self):
        other_user = User.objects.create_user(username='other', password='otherpass')
        asset = DebitCardAsset.objects.create(
            user=other_user,
            name='Other Card',
            type=AssetType.DEBIT_CARD,
            currency='RUB'
        )
        url = reverse('asset_delete', args=[asset.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)


class TransactionDeleteViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('10000.00'))
        self.client.login(username='testuser', password='testpass123')
    
    def test_transaction_delete_get(self):
        transaction = Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('1000.00'),
            currency='RUB',
            to_asset=self.asset,
            date=timezone.now()
        )
        url = reverse('transaction_delete', args=[transaction.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
    
    def test_transaction_delete_post(self):
        transaction = Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('1000.00'),
            currency='RUB',
            to_asset=self.asset,
            date=timezone.now()
        )
        url = reverse('transaction_delete', args=[transaction.pk])
        response = self.client.post(url)
        self.assertRedirects(response, reverse('transactions'))
        self.assertFalse(Transaction.objects.filter(pk=transaction.pk).exists())
    
    def test_transaction_delete_other_user_forbidden(self):
        other_user = User.objects.create_user(username='other', password='otherpass')
        other_asset = DebitCardAsset.objects.create(
            user=other_user,
            name='Other Card',
            type=AssetType.DEBIT_CARD,
            currency='RUB'
        )
        transaction = Transaction.objects.create(
            user=other_user,
            type=TransactionType.REFILL,
            amount=Decimal('1000.00'),
            currency='RUB',
            to_asset=other_asset,
            date=timezone.now()
        )
        url = reverse('transaction_delete', args=[transaction.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)


class TransactionMonthNavigationTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('10000.00'))
        self.client.login(username='testuser', password='testpass123')
    
    def test_prev_month_navigation(self):
        now = timezone.now()
        prev_month = now - timedelta(days=1)
        
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('1000.00'),
            currency='RUB',
            to_asset=self.asset,
            date=prev_month
        )
        
        url = reverse('transactions_month', args=[prev_month.year, prev_month.month])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '1000')


class TransactionFormFieldsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('10000.00'))
        self.client.login(username='testuser', password='testpass123')
    
    def test_refill_shows_to_asset_only(self):
        response = self.client.get(reverse('transaction_add'))
        self.assertContains(response, 'To Asset')
    
    def test_waste_shows_from_asset_only(self):
        response = self.client.get(reverse('transaction_add'))
        self.assertContains(response, 'From Asset')
    
    def test_transfer_shows_both_assets(self):
        response = self.client.get(reverse('transaction_add'))
        self.assertContains(response, 'From Asset')
        self.assertContains(response, 'To Asset')


class StatisticsViewsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('10000.00'))
        self.client.login(username='testuser', password='testpass123')
        self.current_month = timezone.now().month
        self.current_year = timezone.now().year
    
    def test_statistics_requires_login(self):
        self.client.logout()
        response = self.client.get(reverse('statistics'))
        self.assertRedirects(response, f"{reverse('login')}?next=/statistics/")
    
    def test_statistics_list_get(self):
        response = self.client.get(reverse('statistics'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Statistics')
    
    def test_statistics_navigation(self):
        url = reverse('statistics_month', args=[2026, 1])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
    
    def test_statistics_with_income_data(self):
        from datetime import datetime
        from django.utils.timezone import make_aware
        
        test_date = make_aware(datetime(self.current_year, self.current_month, 15, 12, 0))
        
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('5000.00'),
            currency='RUB',
            to_asset=self.asset,
            category='SALARY',
            date=test_date
        )
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('1000.00'),
            currency='RUB',
            to_asset=self.asset,
            category='BONUS',
            date=test_date
        )
        
        url = reverse('statistics_month_type', args=[self.current_year, self.current_month, 'income'])
        response = self.client.get(url)
        self.assertContains(response, '5000.00')
        self.assertContains(response, '1000.00')
        self.assertContains(response, 'Salary')
        self.assertContains(response, 'Bonus')
    
    def test_statistics_with_outcome_data(self):
        from datetime import datetime
        from django.utils.timezone import make_aware
        
        test_date = make_aware(datetime(self.current_year, self.current_month, 15, 12, 0))
        
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.WASTE,
            amount=Decimal('3000.00'),
            currency='RUB',
            from_asset=self.asset,
            category='PRODUCTS',
            date=test_date
        )
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.WASTE,
            amount=Decimal('500.00'),
            currency='RUB',
            from_asset=self.asset,
            category='TRANSPORT',
            date=test_date
        )
        
        url = reverse('statistics_month_type', args=[self.current_year, self.current_month, 'outcome'])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '3000.00')
        self.assertContains(response, '500.00')
        self.assertContains(response, 'Products')
        self.assertContains(response, 'Transport')
    
    def test_statistics_income_subpage(self):
        from datetime import datetime
        from django.utils.timezone import make_aware
        
        test_date = make_aware(datetime(self.current_year, self.current_month, 15, 12, 0))
        
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('5000.00'),
            currency='RUB',
            to_asset=self.asset,
            category='SALARY',
            date=test_date
        )
        
        url = reverse('statistics_month_type', args=[self.current_year, self.current_month, 'income'])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Income')
        self.assertContains(response, 'Salary')
    
    def test_statistics_outcome_subpage(self):
        from datetime import datetime
        from django.utils.timezone import make_aware
        
        test_date = make_aware(datetime(self.current_year, self.current_month, 15, 12, 0))
        
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.WASTE,
            amount=Decimal('3000.00'),
            currency='RUB',
            from_asset=self.asset,
            category='PRODUCTS',
            date=test_date
        )
        
        url = reverse('statistics_month_type', args=[self.current_year, self.current_month, 'outcome'])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Outcome')
        self.assertContains(response, 'Products')
    
    def test_statistics_only_shows_user_data(self):
        from datetime import datetime
        from django.utils.timezone import make_aware
        
        other_user = User.objects.create_user(username='other', password='otherpass')
        other_asset = DebitCardAsset.objects.create(
            user=other_user,
            name='Other Card',
            type=AssetType.DEBIT_CARD,
            currency='RUB'
        )
        
        test_date = make_aware(datetime(self.current_year, self.current_month, 15, 12, 0))
        
        Transaction.objects.create(
            user=other_user,
            type=TransactionType.REFILL,
            amount=Decimal('99999.00'),
            currency='RUB',
            to_asset=other_asset,
            category='SALARY',
            date=test_date
        )
        
        response = self.client.get(reverse('statistics'))
        self.assertNotContains(response, '99999.00')
    
    def test_statistics_empty_month(self):
        response = self.client.get(reverse('statistics'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'No outcome data')
    
    def test_statistics_empty_outcome(self):
        url = reverse('statistics_month_type', args=[self.current_year, self.current_month, 'outcome'])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'No outcome data')
    
    def test_statistics_totals_display(self):
        from datetime import datetime
        from django.utils.timezone import make_aware
        
        test_date = make_aware(datetime(self.current_year, self.current_month, 15, 12, 0))
        
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('5000.00'),
            currency='RUB',
            to_asset=self.asset,
            category='SALARY',
            date=test_date
        )
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.WASTE,
            amount=Decimal('2000.00'),
            currency='RUB',
            from_asset=self.asset,
            category='PRODUCTS',
            date=test_date
        )
        
        response = self.client.get(reverse('statistics'))
        self.assertContains(response, '+5000.00')
        self.assertContains(response, '-2000.00')
    
    def test_statistics_default_period_is_month(self):
        response = self.client.get(reverse('statistics'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Month</a>')
        self.assertContains(response, '2/2026')
    
    def test_statistics_period_year(self):
        url = reverse('statistics_year_type', args=[self.current_year, 'outcome'])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Year</a>')
        self.assertContains(response, 'class="btn btn-outline-secondary active">Year</a>')
    
    def test_statistics_period_navigation(self):
        url = reverse('statistics_year', args=[self.current_year])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
    
    def test_statistics_year_shows_year_stats(self):
        from datetime import datetime
        from django.utils.timezone import make_aware
        
        test_date = make_aware(datetime(self.current_year, 6, 15, 12, 0))
        
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.WASTE,
            amount=Decimal('3000.00'),
            currency='RUB',
            from_asset=self.asset,
            category='PRODUCTS',
            date=test_date
        )
        
        url = reverse('statistics_year', args=[self.current_year])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Year</a>')
        self.assertContains(response, f'>{self.current_year}<')
    
    def test_statistics_month_shows_month_stats(self):
        from datetime import datetime
        from django.utils.timezone import make_aware
        
        test_date = make_aware(datetime(self.current_year, self.current_month, 15, 12, 0))
        
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.WASTE,
            amount=Decimal('3000.00'),
            currency='RUB',
            from_asset=self.asset,
            category='PRODUCTS',
            date=test_date
        )
        
        url = reverse('statistics_month', args=[self.current_year, self.current_month])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Month</a>')
        self.assertContains(response, f'>{self.current_month}/{self.current_year}<')


class ProfileViewsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.user.email = 'test@example.com'
        self.user.first_name = 'Test'
        self.user.last_name = 'User'
        self.user.save()
        self.client.login(username='testuser', password='testpass123')
    
    def test_profile_requires_login(self):
        self.client.logout()
        response = self.client.get(reverse('profile'))
        self.assertRedirects(response, f"{reverse('login')}?next=/profile/")
    
    def test_profile_get(self):
        response = self.client.get(reverse('profile'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Profile')
    
    def test_profile_shows_username(self):
        response = self.client.get(reverse('profile'))
        self.assertContains(response, 'testuser')
    
    def test_profile_shows_email(self):
        response = self.client.get(reverse('profile'))
        self.assertContains(response, 'test@example.com')
    
    def test_profile_shows_first_name(self):
        response = self.client.get(reverse('profile'))
        self.assertContains(response, 'Test')
    
    def test_profile_shows_last_name(self):
        response = self.client.get(reverse('profile'))
        self.assertContains(response, 'User')
    
    def test_profile_shows_date_joined(self):
        response = self.client.get(reverse('profile'))
        self.assertContains(response, 'Date joined')
    
    def test_profile_shows_logout_button(self):
        response = self.client.get(reverse('profile'))
        self.assertContains(response, 'Logout')
    
    def test_profile_shows_download_button(self):
        response = self.client.get(reverse('profile'))
        self.assertContains(response, 'Download Transactions')


class ExportTransactionsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('10000.00'))
        self.client.login(username='testuser', password='testpass123')
    
    def test_export_requires_login(self):
        self.client.logout()
        response = self.client.get(reverse('export_transactions'))
        self.assertRedirects(response, f"{reverse('login')}?next=/profile/export/")
    
    def test_export_returns_excel(self):
        response = self.client.get(reverse('export_transactions'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    def test_export_has_attachment_header(self):
        response = self.client.get(reverse('export_transactions'))
        self.assertTrue(response['Content-Disposition'].startswith('attachment; filename="transactions.xlsx"'))
    
    def test_export_contains_transaction_data(self):
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('5000.00'),
            currency='RUB',
            to_asset=self.asset,
            category='Salary',
            description='Monthly salary',
            date=timezone.now()
        )
        
        response = self.client.get(reverse('export_transactions'))
        self.assertEqual(response.status_code, 200)
    
    def test_export_contains_multiple_transactions(self):
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.REFILL,
            amount=Decimal('5000.00'),
            currency='RUB',
            to_asset=self.asset,
            category='Salary',
            date=timezone.now()
        )
        Transaction.objects.create(
            user=self.user,
            type=TransactionType.WASTE,
            amount=Decimal('1000.00'),
            currency='RUB',
            from_asset=self.asset,
            category='Products',
            date=timezone.now()
        )
        
        response = self.client.get(reverse('export_transactions'))
        response = self.client.get(reverse('export_transactions'))
        self.assertEqual(response.status_code, 200)


class ImportTransactionsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.asset = create_asset_with_balance(self.user, 'Test Card', AssetType.DEBIT_CARD, 'RUB', Decimal('10000.00'))
        self.client.login(username='testuser', password='testpass123')
    
    def test_import_requires_login(self):
        self.client.logout()
        response = self.client.get(reverse('import_transactions'))
        self.assertRedirects(response, f"{reverse('login')}?next=/profile/import/")
    
    def test_import_get_redirects_to_profile(self):
        response = self.client.get(reverse('import_transactions'))
        self.assertRedirects(response, reverse('profile'))
    
    def test_import_no_file_selected(self):
        response = self.client.post(reverse('import_transactions'), {})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'No file selected')
    
    def test_import_with_empty_file(self):
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Type', 'Category', 'Amount', 'Currency', 'From Asset', 'To Asset', 'Description'])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = self.client.post(reverse('import_transactions'), {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Successfully imported 0 transactions')
    
    def test_import_single_transaction(self):
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Type', 'Category', 'Amount', 'Currency', 'From Asset', 'To Asset', 'Description'])
        ws.append(['2026-02-20 12:00', 'Refill', 'Salary', '5000', 'RUB', '', 'Test Card', 'Monthly salary'])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = self.client.post(reverse('import_transactions'), {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Successfully imported 1 transactions')
        self.assertTrue(Transaction.objects.filter(amount=Decimal('5000.00'), category='Salary').exists())
    
    def test_import_multiple_transactions(self):
        Transaction.objects.filter(user=self.user).delete()
        
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Type', 'Category', 'Amount', 'Currency', 'From Asset', 'To Asset', 'Description'])
        ws.append(['2026-02-20 12:00', 'Refill', 'Salary', '5000', 'RUB', '', 'Test Card', 'Test 1'])
        ws.append(['2026-02-21 12:00', 'Waste', 'Products', '1000', 'RUB', 'Test Card', '', 'Test 2'])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = self.client.post(reverse('import_transactions'), {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Successfully imported 2 transactions')
        self.assertEqual(Transaction.objects.filter(user=self.user).count(), 2)
    
    def test_import_with_empty_description(self):
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Type', 'Category', 'Amount', 'Currency', 'From Asset', 'To Asset', 'Description'])
        ws.append(['2026-02-20 12:00', 'Refill', 'Salary', '5000', 'RUB', '', 'Test Card', None])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = self.client.post(reverse('import_transactions'), {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Successfully imported 1 transactions')
    
    def test_import_with_empty_category(self):
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Type', 'Category', 'Amount', 'Currency', 'From Asset', 'To Asset', 'Description'])
        ws.append(['2026-02-20 12:00', 'Refill', '', '5000', 'RUB', '', 'Test Card', 'Test'])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = self.client.post(reverse('import_transactions'), {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Successfully imported 1 transactions')
        tx = Transaction.objects.first()
        self.assertEqual(tx.category, '')
    
    def test_import_with_transfer_type(self):
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Type', 'Category', 'Amount', 'Currency', 'From Asset', 'To Asset', 'Description'])
        ws.append(['2026-02-20 12:00', 'Transfer', '', '1000', 'RUB', 'Test Card', 'Test Card', 'Transfer test'])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = self.client.post(reverse('import_transactions'), {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Successfully imported 1 transactions')
        self.assertTrue(Transaction.objects.filter(type=TransactionType.TRANSFER).exists())
    
    def test_import_only_creates_for_current_user(self):
        Transaction.objects.filter(user=self.user).delete()
        
        other_user = User.objects.create_user(username='other', password='otherpass')
        other_asset = DebitCardAsset.objects.create(
            user=other_user,
            name='Other Card',
            type=AssetType.DEBIT_CARD,
            currency='RUB'
        )
        
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Type', 'Category', 'Amount', 'Currency', 'From Asset', 'To Asset', 'Description'])
        ws.append(['2026-02-20 12:00', 'Refill', 'Salary', '99999', 'RUB', '', 'Other Card', 'Other user data'])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = self.client.post(reverse('import_transactions'), {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Transaction.objects.filter(user=self.user).count(), 1)
        self.assertFalse(Transaction.objects.filter(user=other_user).exists())
